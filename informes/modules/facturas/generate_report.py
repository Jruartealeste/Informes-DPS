"""
Genera el informe dinamico de Facturas a partir de lo que haya en la base
local en este momento. Se reconstruye cada vez que se corre, reflejando la
ultima carga hecha con ingest.py.

Uso:
    python -m modules.facturas.generate_report
"""
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

import db
from . import config


def cargar_datos() -> pd.DataFrame:
    with db.get_connection() as conn:
        df = pd.read_sql_query(f"SELECT * FROM {config.DB_TABLE}", conn)
    if not df.empty:
        df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
        df["mes"] = df["fecha"].dt.to_period("M").astype(str)
    return df


def escribir_hoja(ws, df: pd.DataFrame, titulo: str):
    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for cell in ws[3]:
        cell.font = Font(bold=True)
    for col_cells in ws.columns:
        largo = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(largo + 2, 40)


def main():
    df = cargar_datos()
    if df.empty:
        print("No hay datos cargados todavia. Corre 'python -m modules.facturas.ingest' primero.")
        return

    resumen_cliente = (
        df.groupby("cliente")
        .agg(
            facturas=("clave_factura", "count"),
            subtotal_ml=("subtotal_ml", "sum"),
            total_ml=("total_ml", "sum"),
        )
        .reset_index()
        .sort_values("total_ml", ascending=False)
    )

    resumen_mes = (
        df.groupby("mes")
        .agg(
            facturas=("clave_factura", "count"),
            subtotal_ml=("subtotal_ml", "sum"),
            total_ml=("total_ml", "sum"),
        )
        .reset_index()
        .sort_values("mes")
    )

    resumen_producto = (
        df.groupby("producto")
        .agg(facturas=("clave_factura", "count"), total_ml=("total_ml", "sum"))
        .reset_index()
        .sort_values("total_ml", ascending=False)
    )

    resumen_anunciante = (
        df.groupby("anunciante")
        .agg(facturas=("clave_factura", "count"), total_ml=("total_ml", "sum"))
        .reset_index()
        .sort_values("total_ml", ascending=False)
    )

    wb = Workbook()

    ws0 = wb.active
    ws0.title = "Resumen"
    ws0.append(["Informe de Facturas - Advertys"])
    ws0["A1"].font = Font(bold=True, size=14)
    ws0.append([f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws0.append([f"Facturas totales: {len(df)}"])
    ws0.append([f"Clientes distintos: {df['cliente'].nunique()}"])
    ws0.append([f"Anunciantes distintos: {df['anunciante'].nunique()}"])
    ws0.append([f"Total ML facturado: {df['total_ml'].sum():,.2f}"])
    ws0.append([f"Subtotal ML facturado: {df['subtotal_ml'].sum():,.2f}"])

    ws1 = wb.create_sheet("Por Cliente")
    escribir_hoja(ws1, resumen_cliente, "Facturacion por cliente")

    ws2 = wb.create_sheet("Por Mes")
    escribir_hoja(ws2, resumen_mes, "Facturacion por mes (segun fecha de factura)")

    chart = BarChart()
    chart.title = "Total ML facturado por mes"
    chart.y_axis.title = "Total ML"
    chart.x_axis.title = "Mes"
    data = Reference(ws2, min_col=4, min_row=3, max_row=3 + len(resumen_mes))
    cats = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(resumen_mes))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "G3")

    ws3 = wb.create_sheet("Por Anunciante")
    escribir_hoja(ws3, resumen_anunciante, "Facturacion por anunciante (marca)")

    ws4 = wb.create_sheet("Por Producto")
    escribir_hoja(ws4, resumen_producto, "Facturacion por producto")

    ws5 = wb.create_sheet("Detalle")
    escribir_hoja(ws5, df.drop(columns=["mes"]), "Detalle completo de facturas")

    wb.save(config.REPORT_OUTPUT_PATH)
    print(f"OK: informe generado en {config.REPORT_OUTPUT_PATH}")


if __name__ == "__main__":
    main()
