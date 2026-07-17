"""
Genera el informe dinamico de Ordenes de Trabajo a partir de lo que haya en
la base local en este momento. Cada vez que se corre, refleja los datos mas
recientes que se hayan cargado con ingest.py (por eso es "dinamico": no es
un Excel fijo, se reconstruye).

Uso:
    python -m modules.ordenes_trabajo.generate_report
"""
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

import db
from . import config


def cargar_datos() -> pd.DataFrame:
    with db.get_connection() as conn:
        df = pd.read_sql_query(f"SELECT * FROM {config.DB_TABLE}", conn)
    if not df.empty:
        df["fecha_abierta"] = pd.to_datetime(df["fecha_abierta"], errors="coerce")
        df["mes"] = df["fecha_abierta"].dt.to_period("M").astype(str)
    return df


def escribir_hoja(ws, df: pd.DataFrame, titulo: str):
    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for cell in ws[3]:
        cell.font = Font(bold=True)
    for col_cells in ws.columns:
        largo = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(largo + 2, 40)


def main():
    df = cargar_datos()
    if df.empty:
        print("No hay datos cargados todavia. Corre 'python -m modules.ordenes_trabajo.ingest' primero.")
        return

    resumen_anunciante = (
        df.groupby("anunciante")
        .agg(
            ordenes=("numero_ot", "count"),
            renta_teorica=("renta_teorica", "sum"),
            renta_real=("renta_real", "sum"),
        )
        .reset_index()
        .sort_values("renta_real", ascending=False)
    )

    resumen_mes = (
        df.groupby("mes")
        .agg(
            ordenes=("numero_ot", "count"),
            renta_teorica=("renta_teorica", "sum"),
            renta_real=("renta_real", "sum"),
        )
        .reset_index()
        .sort_values("mes")
    )

    resumen_estado = (
        df.groupby("estado")
        .agg(ordenes=("numero_ot", "count"), renta_real=("renta_real", "sum"))
        .reset_index()
        .sort_values("ordenes", ascending=False)
    )

    wb = Workbook()

    ws0 = wb.active
    ws0.title = "Resumen"
    ws0.append(["Informe dinamico - Advertys"])
    ws0["A1"].font = Font(bold=True, size=14)
    ws0.append([f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws0.append([f"Ordenes de trabajo totales: {len(df)}"])
    ws0.append([f"Renta real total: {df['renta_real'].sum():,.2f}"])
    ws0.append([f"Renta teorica total: {df['renta_teorica'].sum():,.2f}"])

    ws1 = wb.create_sheet("Por Anunciante")
    escribir_hoja(ws1, resumen_anunciante, "Rentabilidad por anunciante")

    ws2 = wb.create_sheet("Por Mes")
    escribir_hoja(ws2, resumen_mes, "Rentabilidad por mes")

    chart = BarChart()
    chart.title = "Renta real por mes"
    chart.y_axis.title = "Renta real"
    chart.x_axis.title = "Mes"
    data = Reference(ws2, min_col=4, min_row=3, max_row=3 + len(resumen_mes))
    cats = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(resumen_mes))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "G3")

    ws3 = wb.create_sheet("Por Estado")
    escribir_hoja(ws3, resumen_estado, "Ordenes por estado")

    ws4 = wb.create_sheet("Detalle")
    escribir_hoja(ws4, df.drop(columns=["mes"]), "Detalle completo de ordenes de trabajo")

    wb.save(config.REPORT_OUTPUT_PATH)
    print(f"OK: informe generado en {config.REPORT_OUTPUT_PATH}")


if __name__ == "__main__":
    main()
