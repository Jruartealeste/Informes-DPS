"""
Genera el informe dinamico de Compras a partir de lo que haya en la base
local en este momento. Se reconstruye cada vez que se corre, reflejando la
ultima carga hecha con ingest.py.

Uso:
    python -m modules.compras.generate_report
"""
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

import db
from . import config


def cargar_datos() -> pd.DataFrame:
    with db.get_connection() as conn:
        df = pd.read_sql_query(f"SELECT * FROM {config.DB_TABLE}", conn)
    if not df.empty:
        df["fecha_factura"] = pd.to_datetime(df["fecha_factura"], errors="coerce")
        df["mes"] = df["fecha_factura"].dt.to_period("M").astype(str)
    return df


def escribir_hoja(ws, df: pd.DataFrame, titulo: str):
    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for cell in ws[3]:
        cell.font = Font(bold=True)
    for col_cells in ws.columns:
        largo = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(largo + 2, 40)


def main():
    df = cargar_datos()
    if df.empty:
        print("No hay datos cargados todavia. Corre 'python -m modules.compras.ingest' primero.")
        return

    resumen_proveedor = (
        df.groupby("proveedor")
        .agg(
            compras=("clave_compra", "count"),
            importe_sin_iva=("importe_sin_iva_signado", "sum"),
            total_impositivo=("total_impositivo", "sum"),
        )
        .reset_index()
        .sort_values("total_impositivo", ascending=False)
    )

    resumen_mes = (
        df.groupby("mes")
        .agg(
            compras=("clave_compra", "count"),
            importe_sin_iva=("importe_sin_iva_signado", "sum"),
            total_impositivo=("total_impositivo", "sum"),
        )
        .reset_index()
        .sort_values("mes")
    )

    resumen_tipo = (
        df.groupby("tipo_compra")
        .agg(compras=("clave_compra", "count"), total_impositivo=("total_impositivo", "sum"))
        .reset_index()
        .sort_values("total_impositivo", ascending=False)
    )

    resumen_estado = (
        df.groupby("estado")
        .agg(compras=("clave_compra", "count"), total_impositivo=("total_impositivo", "sum"))
        .reset_index()
        .sort_values("compras", ascending=False)
    )

    contabilizadas = df[df["estado"] == "Contabilizado"]

    wb = Workbook()

    ws0 = wb.active
    ws0.title = "Resumen"
    ws0.append(["Informe de Compras - Advertys"])
    ws0["A1"].font = Font(bold=True, size=14)
    ws0.append([f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws0.append([f"Compras totales (todos los estados): {len(df)}"])
    ws0.append([f"Compras contabilizadas: {len(contabilizadas)}"])
    ws0.append([f"Proveedores distintos: {df['proveedor'].nunique()}"])
    ws0.append([f"Total impositivo (contabilizadas): {contabilizadas['total_impositivo'].sum():,.2f}"])
    ws0.append([f"Importe s/IVA con signo (contabilizadas): {contabilizadas['importe_sin_iva_signado'].sum():,.2f}"])

    ws1 = wb.create_sheet("Por Proveedor")
    escribir_hoja(ws1, resumen_proveedor, "Compras por proveedor")

    ws2 = wb.create_sheet("Por Mes")
    escribir_hoja(ws2, resumen_mes, "Compras por mes (segun fecha de factura)")

    chart = BarChart()
    chart.title = "Total impositivo por mes"
    chart.y_axis.title = "Total impositivo"
    chart.x_axis.title = "Mes"
    data = Reference(ws2, min_col=4, min_row=3, max_row=3 + len(resumen_mes))
    cats = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(resumen_mes))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "G3")

    ws3 = wb.create_sheet("Por Tipo de Compra")
    escribir_hoja(ws3, resumen_tipo, "Compras por tipo (Gastos / Medios / Produccion)")

    ws4 = wb.create_sheet("Por Estado")
    escribir_hoja(ws4, resumen_estado, "Compras por estado")

    ws5 = wb.create_sheet("Detalle")
    escribir_hoja(ws5, df.drop(columns=["mes"]), "Detalle completo de compras")

    wb.save(config.REPORT_OUTPUT_PATH)
    print(f"OK: informe generado en {config.REPORT_OUTPUT_PATH}")


if __name__ == "__main__":
    main()
